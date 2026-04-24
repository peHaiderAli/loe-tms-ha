import json
import math
import re
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


def normalize_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")


def normalize_lookup(value: str | None) -> str:
    return re.sub(r"[^a-z0-9]+", "", (value or "").lower())


def serializable(value):
    if value in (None, "", "-"):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float):
        if math.isnan(value):
            return None
        return round(value, 2)
    return value


def infer_type(name: str) -> str:
    if "pixeledge: platform" in name.lower():
        return "Internal platform"
    if "pixeledge: processes" in name.lower():
        return "Internal operations"
    if "pixeledge" in name.lower() or name.lower() == "other":
        return "Internal support"
    if "fintech" in name.lower() or "loanedge" in name.lower() or "bdc" in name.lower():
        return "Product"
    return "Client delivery"


def infer_priority(name: str) -> str:
    lowered = name.lower()
    if any(token in lowered for token in ["erm", "fintech", "loanedge", "bdc", "hsf"]):
        return "A"
    if "pixeledge" in lowered:
        return "B"
    return "C"


def owner_name(primary, secondary):
    primary = serializable(primary)
    secondary = serializable(secondary)
    if primary and secondary and primary != secondary:
        return f"{primary} / {secondary}"
    return primary or secondary or "Unassigned"


def load_priority_map(workbook):
    if "Projects & Priorities" not in workbook.sheetnames:
        return {}

    ws = workbook["Projects & Priorities"]
    priorities = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        priority, type_name, project_name, owner = row[:4]
        if not project_name:
            continue
        priorities[normalize_lookup(project_name)] = {
            "priority": serializable(priority) or infer_priority(str(project_name)),
            "type": serializable(type_name) or infer_type(str(project_name)),
            "owner_name": serializable(owner) or "Unassigned",
        }
    return priorities


def parse(path: Path):
    workbook = load_workbook(path, data_only=True)
    live = workbook["Live Allocation Guidelines"]
    priority_map = load_priority_map(workbook)

    row1 = [cell.value for cell in live[1]]
    row2 = [cell.value for cell in live[2]]
    headers = [cell.value for cell in live[3]]

    project_columns = []
    project_totals = defaultdict(float)

    for index in range(14, len(headers)):
        name = headers[index]
        if not name:
            continue
        normalized = normalize_lookup(str(name))
        priority_data = priority_map.get(normalized, {})
        project_columns.append(
            {
                "index": index,
                "name": str(name),
                "project_key": normalize_key(str(name)),
                "priority": priority_data.get("priority", infer_priority(str(name))),
                "type": priority_data.get("type", infer_type(str(name))),
                "owner_name": priority_data.get("owner_name", owner_name(row1[index] if index < len(row1) else None, row2[index] if index < len(row2) else None)),
            }
        )

    employees = []
    for row in live.iter_rows(min_row=4, values_only=True):
        if not row or not row[0]:
            continue

        name = serializable(row[0])
        if not name or name == "#REF!":
            continue

        allocations = []
        for project in project_columns:
            value = row[project["index"]] if project["index"] < len(row) else None
            effort = serializable(value)
            if isinstance(effort, (int, float)) and effort > 0:
                allocations.append(
                    {
                        "project_key": project["project_key"],
                        "effort": effort,
                        "note": "Imported from Live Allocation Guidelines",
                    }
                )
                project_totals[project["project_key"]] += effort

        employees.append(
            {
                "source_key": normalize_key(str(name)),
                "name": name,
                "preferred_name": serializable(row[1]),
                "stream": serializable(row[2]),
                "skills": serializable(row[3]),
                "title": serializable(row[4]),
                "location": serializable(row[5]),
                "start_date": serializable(row[6]),
                "onboarding_mentor": serializable(row[7]),
                "growth_partner": serializable(row[8]),
                "reviewer_name": serializable(row[9]),
                "review_status": serializable(row[10]),
                "email": None,
                "allocations": allocations,
            }
        )

    projects = []
    for project in project_columns:
        projects.append(
            {
                "name": project["name"],
                "project_key": project["project_key"],
                "priority": project["priority"],
                "type": project["type"],
                "owner_name": project["owner_name"],
                "target_capacity": max(project_totals.get(project["project_key"], 0.0), 1.0),
            }
        )

    return {"projects": projects, "employees": employees}


if __name__ == "__main__":
    workbook_path = Path(sys.argv[1])
    print(json.dumps(parse(workbook_path), indent=2))
